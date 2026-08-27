from collections import Counter, defaultdict
from hashlib import sha256
from pathlib import Path
import json
import re
import zipfile

from openpyxl import load_workbook


ROOT = Path(".")
MAPPING = ROOT / "outputs" / "Table S3_R108_mapping_QC.xlsx"
S4 = ROOT / "NCC1 supplemental tables" / "Table S4.xlsx"
R108_FASTA = ROOT / "medtr.R108.gnmHiC_1.ann1.Y8NH.protein.faa"
WORK = ROOT / "work" / "figure2"
OUTPUTS = ROOT / "outputs"
FASTA_OUT = OUTPUTS / "Figure2_110_unique_R108_ColabFold_input.fasta"
BATCH_DIR = OUTPUTS / "Figure2_ColabFold_batches"
INDIVIDUAL_DIR = OUTPUTS / "Figure2_ColabFold_89_short_individual_fastas"
INDIVIDUAL_ZIP = OUTPUTS / "Figure2_ColabFold_89_short_individual_fastas.zip"
JSON_OUT = WORK / "Figure2_candidate_manifest.json"
QC_OUT = OUTPUTS / "Figure2_candidate_sequence_QC.txt"

USABLE = {"Exact unique", "High confidence", "Probable"}


def read_table(path, sheet_name, header_row):
    ws = load_workbook(path, read_only=True, data_only=True)[sheet_name]
    it = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(it)
    headers = list(next(it))
    return [dict(zip(headers, row)) for row in it if any(v is not None for v in row)]


def read_fasta(path):
    records = {}
    current = None
    chunks = []
    with path.open() as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if current is not None:
                    records[current] = "".join(chunks).upper()
                raw_id = line[1:].split()[0]
                match = re.search(r"(MtrunR108HiC_\d+\.\d+)$", raw_id)
                current = match.group(1) if match else raw_id
                chunks = []
            else:
                chunks.append(line)
    if current is not None:
        records[current] = "".join(chunks).upper()
    return records


WORK.mkdir(parents=True, exist_ok=True)
OUTPUTS.mkdir(parents=True, exist_ok=True)
BATCH_DIR.mkdir(parents=True, exist_ok=True)
INDIVIDUAL_DIR.mkdir(parents=True, exist_ok=True)

mapping_rows = read_table(MAPPING, "Mapping QC", 1)
s4_rows = read_table(S4, "Hoja1", 2)
r108_sequences = read_fasta(R108_FASTA)
s4_accessions = {row["Uniprot Accession"] for row in s4_rows if row["Uniprot Accession"]}

selected = [
    row for row in mapping_rows
    if row["UniProt accession"] in s4_accessions and row["Mapping status"] in USABLE
]
selected.sort(key=lambda row: (row["R108 protein ID"], row["UniProt accession"]))

by_r108 = defaultdict(list)
for row in selected:
    by_r108[row["R108 protein ID"]].append(row)

manifest = []
missing = []
for r108_id, rows in sorted(by_r108.items()):
    sequence = r108_sequences.get(r108_id)
    if not sequence:
        missing.append(r108_id)
        continue
    accessions = sorted(row["UniProt accession"] for row in rows)
    statuses = sorted({row["Mapping status"] for row in rows})
    annotations = sorted({str(row["R108 annotation"]) for row in rows if row["R108 annotation"]})
    cxxc_positions = [m.start() + 1 for m in re.finditer(r"C..C", sequence)]
    manifest.append({
        "prediction_id": r108_id,
        "r108_protein_id": r108_id,
        "r108_gene_id": rows[0]["R108 gene ID"],
        "uniprot_accessions": ";".join(accessions),
        "a17_gene_ids": ";".join(sorted({str(row["A17 gene ID"]) for row in rows if row["A17 gene ID"]})),
        "mapping_statuses": ";".join(statuses),
        "best_identity_pct": max(float(row["Identity (%)"]) for row in rows),
        "best_coverage_pct": max(float(row["Coverage (%)"]) for row in rows),
        "min_identity_margin_pp": min(float(row["Identity margin (pp)"]) for row in rows),
        "r108_annotation": "; ".join(annotations),
        "sequence_length": len(sequence),
        "cysteine_count": sequence.count("C"),
        "cxxc_count": len(cxxc_positions),
        "cxxc_positions_1based": ",".join(map(str, cxxc_positions)),
        "sequence_sha256": sha256(sequence.encode()).hexdigest(),
        "source_table_s4": "Yes",
        "prediction_status": "Not run",
        "structure_file": "",
        "mean_plddt": "",
        "high_confidence_fraction_plddt70": "",
        "candidate_class": "Pending structure",
        "manual_review": "Pending",
        "cys_pair_screen_eligible": "Yes" if sequence.count("C") >= 2 else "No",
        "exclusion_reason": "" if sequence.count("C") >= 2 else "Fewer than two cysteines",
        "compute_batch": (
            "01_le1000aa" if len(sequence) <= 1000
            else "02_1001_1500aa" if len(sequence) <= 1500
            else "03_gt1500aa_individual"
        ),
        "sequence": sequence,
    })

hash_to_ids = defaultdict(list)
for row in manifest:
    hash_to_ids[row["sequence_sha256"]].append(row["r108_protein_id"])
duplicate_sequence_groups = [ids for ids in hash_to_ids.values() if len(ids) > 1]

with FASTA_OUT.open("w") as handle:
    for row in manifest:
        handle.write(f">{row['prediction_id']}\n")
        sequence = row["sequence"]
        for start in range(0, len(sequence), 80):
            handle.write(sequence[start:start + 80] + "\n")

batch_files = {
    "01_le1000aa": BATCH_DIR / "Figure2_CysPair_targets_01_le1000aa.fasta",
    "02_1001_1500aa": BATCH_DIR / "Figure2_CysPair_targets_02_1001_1500aa.fasta",
    "03_gt1500aa_individual": BATCH_DIR / "Figure2_CysPair_targets_03_gt1500aa.fasta",
}
for batch_name, batch_path in batch_files.items():
    with batch_path.open("w") as handle:
        for row in manifest:
            if row["cys_pair_screen_eligible"] != "Yes" or row["compute_batch"] != batch_name:
                continue
            handle.write(f">{row['prediction_id']}\n")
            sequence = row["sequence"]
            for start in range(0, len(sequence), 80):
                handle.write(sequence[start:start + 80] + "\n")

short_rows = [
    row for row in manifest
    if row["cys_pair_screen_eligible"] == "Yes" and row["compute_batch"] == "01_le1000aa"
]
for row in short_rows:
    target = INDIVIDUAL_DIR / f"{row['prediction_id']}.fasta"
    with target.open("w") as handle:
        handle.write(f">{row['prediction_id']}\n")
        sequence = row["sequence"]
        for start in range(0, len(sequence), 80):
            handle.write(sequence[start:start + 80] + "\n")
with zipfile.ZipFile(INDIVIDUAL_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as archive:
    for target in sorted(INDIVIDUAL_DIR.glob("*.fasta")):
        archive.write(target, arcname=f"input_fasta/{target.name}")

JSON_OUT.write_text(json.dumps({
    "summary": {
        "table_s4_unique_accessions": len(s4_accessions),
        "usable_accession_mappings": len(selected),
        "unique_r108_prediction_ids": len(manifest),
        "missing_r108_sequences": len(missing),
        "r108_ids_with_multiple_a17_accessions": sum(len(rows) > 1 for rows in by_r108.values()),
        "duplicate_sequence_groups_across_r108_ids": len(duplicate_sequence_groups),
        "cxxc_positive_r108_sequences": sum(row["cxxc_count"] > 0 for row in manifest),
        "cxxc_negative_r108_sequences": sum(row["cxxc_count"] == 0 for row in manifest),
        "zero_cysteine_sequences": sum(row["cysteine_count"] == 0 for row in manifest),
        "one_cysteine_sequences": sum(row["cysteine_count"] == 1 for row in manifest),
        "cys_pair_screen_eligible_unique_sequences": sum(row["cysteine_count"] >= 2 for row in manifest),
        "length_le_1000": sum(row["sequence_length"] <= 1000 for row in manifest),
        "length_1001_1500": sum(1000 < row["sequence_length"] <= 1500 for row in manifest),
        "length_gt_1500": sum(row["sequence_length"] > 1500 for row in manifest),
        "max_sequence_length": max(row["sequence_length"] for row in manifest),
        "median_sequence_length": sorted(row["sequence_length"] for row in manifest)[len(manifest) // 2],
    },
    "missing_r108_ids": missing,
    "duplicate_sequence_groups": duplicate_sequence_groups,
    "records": manifest,
}, indent=2, ensure_ascii=False) + "\n")

summary = json.loads(JSON_OUT.read_text())["summary"]
qc_lines = [
    "Figure 2 structure-prediction input QC",
    "",
    f"Table S4 unique Medicago accessions: {summary['table_s4_unique_accessions']}",
    f"Usable accession-level A17-to-R108 mappings: {summary['usable_accession_mappings']}",
    f"Unique R108 prediction IDs written to FASTA: {summary['unique_r108_prediction_ids']}",
    f"Missing R108 sequences: {summary['missing_r108_sequences']}",
    f"R108 IDs linked to multiple A17 accessions: {summary['r108_ids_with_multiple_a17_accessions']}",
    f"Duplicate sequence groups across different R108 IDs: {summary['duplicate_sequence_groups_across_r108_ids']}",
    f"CXXC-positive R108 sequences: {summary['cxxc_positive_r108_sequences']}",
    f"CXXC-negative R108 sequences: {summary['cxxc_negative_r108_sequences']}",
    f"Sequences with zero cysteine: {summary['zero_cysteine_sequences']}",
    f"Sequences with one cysteine: {summary['one_cysteine_sequences']}",
    f"Unique sequences eligible for a Cys-pair screen (>=2 Cys): {summary['cys_pair_screen_eligible_unique_sequences']}",
    "",
    "Length bins for compute planning",
    f"<=1000 aa: {summary['length_le_1000']}",
    f"1001-1500 aa: {summary['length_1001_1500']}",
    f">1500 aa: {summary['length_gt_1500']}",
    f"Median length: {summary['median_sequence_length']} aa",
    f"Maximum length: {summary['max_sequence_length']} aa",
    "",
    "Selection rule",
    "Table S4 accession AND mapping status in {Exact unique, High confidence, Probable}.",
    "One record is emitted per unique R108 protein ID; accession-level provenance is retained in the manifest workbook.",
    "The all-sequence FASTA contains 110 unique R108 proteins. Compute batches contain only proteins with at least two cysteines.",
]
QC_OUT.write_text("\n".join(qc_lines) + "\n")
print(json.dumps(summary, indent=2))
